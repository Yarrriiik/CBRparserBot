import json
from openpyxl import load_workbook


# def parse_xlsx_to_json(path):
#     with open('result.json', 'rt', encoding='utf-8') as file:
#         data_dict = json.loads(file.read())
#         print(data_dict)
#     wb = load_workbook(path)
#     ws = wb['RC']
#     data_dict = {}
#     for i in range(1, ws.max_row):
#         data_dict[ws[f'B{i}'].value] = [ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value]
#     print(data_dict)
#     with open('result.json', 'w', encoding='utf-8') as fp:
#         json.dump(data_dict, fp, indent=4, ensure_ascii=False)
# parse_xlsx_to_json('XLUnlegal.xlsx')


def parse_xlsx_to_json(path):
    wb = load_workbook(path)
    ws = wb['RC']
    data_dict = {}
    for i in range(1, ws.max_row):
        if ws[f'B{i}'].value is not None:
            data_dict[ws[f'B{i}'].value] = [ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value]
        if ws[f'C{i}'].value is not None:
            data_dict[ws[f'C{i}'].value] = [ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value]
        if ws[f'D{i}'].value is not None:
            data_dict[ws[f'D{i}'].value] = [ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value]
        if ws[f'E{i}'].value is not None:
            data_dict[ws[f'E{i}'].value] = [ws[f'B{i}'].value, ws[f'C{i}'].value, ws[f'D{i}'].value, ws[f'E{i}'].value, ws[f'F{i}'].value, ws[f'G{i}'].value, ws[f'H{i}'].value]
    with open('result.json', 'w', encoding='utf-8') as fp:
        json.dump(data_dict, fp, indent=4, ensure_ascii=False)
parse_xlsx_to_json('XLUnlegal.xlsx')




# dict1 = {'AAPL': ['Apple Inc.']}
#
# sett = set(dict1.keys())
# print(sett)
# flag = []
#
# for i in sett:
#     if '' in i:
#         flag.append(i)
#
# for i in flag:
#     print(dict1[i])


@dp.message_handler(commands='reply_xlsx_file', state='*')
async def reply_xlsx_files(message: types.Message, state: FSMContext):
    await state.set_state(rezult_xlsx.update_file.state)
@dp.message_handler(content_types=ContentTypes.DOCUMENT)
async def unknown_message(message: types.Message):
    file_path = "path/to/XLUnlegal.xlsx"
    if os.path.exists(file_path) == True:
        os.remove('XLUnlegal.xlsx')
    if document := message.document:
        if document.file_name == 'XLUnlegal.xlsx':
            await document.download(destination_file=f'{document.file_name}')
            await message.answer('Файл обновлён.')
        else:
            await message.answer('Не правильное название файла. Необходимо: "XLUnlegal.xlsx".')