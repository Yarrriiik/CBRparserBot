import logging
import os
import warnings
import aiogram
import openpyxl
import json
from kbluid import *
from openpyxl import load_workbook
from aiogram import types, Bot, Dispatcher, executor
from aiogram.dispatcher.filters import Text
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, KeyboardButton, ReplyKeyboardMarkup, message, \
    ContentTypes
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from datetime import datetime


bot = Bot(token="6294783042:AAF6Y2AknMd7qGb_UMQzdaaHZmEesxyPl-0")
dp = Dispatcher(bot, storage=MemoryStorage())
logging.basicConfig(level=logging.INFO)
warnings.simplefilter("ignore")


class processing_class(StatesGroup):
    check = State()
    rs = State()
    result = State()

class rezult_xlsx(StatesGroup):
    update_file = State()


def parse_xlsx_to_json(path):
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    data_dict = {}
    for i in range(1, ws.max_row+1):
        try:
            if ws[f'B{i}'].value is not None:
                data_dict[ws[f'B{i}'].value.replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО')] = [ws[f'B{i}'].value.replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО'), ws[f'C{i}'].value, ws[f'D{i}'].value,
                                                ws[f'E{i}'].value, ws[f'F{i}'].value,
                                                'Дата' if ws[f'A{i}'].value == 'Дата' else ws[f'A{i}'].value.strftime(
                                                    '%d.%m.%y')]
            if ws[f'C{i}'].value is not None:
                data_dict[ws[f'C{i}'].value] = [ws[f'B{i}'].value.replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО'), ws[f'C{i}'].value, ws[f'D{i}'].value,
                                                ws[f'E{i}'].value, ws[f'F{i}'].value,
                                                'Дата' if ws[f'A{i}'].value == 'Дата' else ws[f'A{i}'].value.strftime('%d.%m.%y')]
            if ws[f'D{i}'].value is not None:
                data_dict[ws[f'D{i}'].value] = [ws[f'B{i}'].value.replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО'), ws[f'C{i}'].value, ws[f'D{i}'].value,
                                                ws[f'E{i}'].value, ws[f'F{i}'].value,
                                                'Дата' if ws[f'A{i}'].value == 'Дата' else ws[f'A{i}'].value.strftime('%d.%m.%y')]
            if ws[f'E{i}'].value is not None:
                data_dict[ws[f'E{i}'].value] = [ws[f'B{i}'].value.replace('ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ', 'ООО'), ws[f'C{i}'].value, ws[f'D{i}'].value,
                                                ws[f'E{i}'].value, ws[f'F{i}'].value,
                                                'Дата' if ws[f'A{i}'].value == 'Дата' else ws[f'A{i}'].value.strftime('%d.%m.%y')]
        except ValueError:
            pass
    with open('result.json', 'w', encoding='utf-8') as fp:
        json.dump(data_dict, fp, indent=4, ensure_ascii=False)

@dp.message_handler(commands='reply_xlsx_file', state='*')
async def reply_xlsx_files(message: types.Message, state: FSMContext):
    await state.set_state(rezult_xlsx.update_file.state)
    await message.answer('Отправьте файл')
@dp.message_handler(content_types=ContentTypes.DOCUMENT, state=rezult_xlsx.update_file.state)
async def unknown_message(message: types.Message):
    try:
        os.remove('result.json')
    except FileNotFoundError:
        pass
    if document := message.document:
        await document.download(destination_file=f'{document.file_name}')
        parse_xlsx_to_json(path=f'{document.file_name}')
        await message.answer('Файл обновлён.')
        os.remove(f'{document.file_name}')

@dp.callback_query_handler(Text(equals='selection_menu'), state='*')
async def selection_menu_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await call.message.delete()
    await bot.send_message(chat_id=call.from_user.id,
                           text='📖Ознакомьтесь с основными видами нелегальной деятельности на финансовом рынке',
                           reply_markup=osmenu)
    await bot.send_message(chat_id=call.from_user.id,
                           text='💰Быстрый заработок, инвестиционные проекты, пассивный доход:',
                           parse_mode='HTML',
                           reply_markup=menu1)
    await bot.send_message(chat_id=call.from_user.id,
                           text='💳 Кредитование, выдача займов, займы под залог ПТС / недвижимости:',
                           parse_mode='HTML',
                           reply_markup=menu2)
    await bot.send_message(chat_id=call.from_user.id,
                           text='📈 Брокер, форекс-дилер, доверительное управление (профессиональные участники рынка ценных бумаг): рынка ценных бумаг):',
                           parse_mode='HTML',
                           reply_markup=menu8)
    await bot.send_message(chat_id=call.from_user.id,
                           text='📄 Страхование, полисы ОСАГО::',
                           reply_markup=menu11)
    await state.finish()
@dp.message_handler(text='↩️Возврат в меню выбора', state='*')
async def cancel_command(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer(text='Для перехода в меню:',
                        reply_markup=tran3)

@dp.message_handler(text='🔎Справочник финансовых организаций', state='*')
async def sprav_command(message: types. Message, state: FSMContext):
    await state.finish()
    await message.answer(text='Проверить финансовую организацию:',
                        reply_markup=menu4)

@dp.message_handler(text='🆘Warning list Банка России', state='*')
async def cancel_command(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer('В справочнике нелегальной деятельности организаций можно узнать информации по: "Поиск по названию, ИНН, адресу, сайту за весь доступный период".',
                           reply_markup=menu7)

@dp.message_handler(text='📝Интернет-приёмная', state='*')
async def cancel_command(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer('Перейти в Интернет-приёмную Банка России:',
                         reply_markup=tran1)
@dp.message_handler(text='☎️Контактная информация', state='*')
async def cancel_command(message: types.Message, state: FSMContext):
    await state.finish()
    await message.answer('Связаться с Банком России:',
                         reply_markup=tran2)


@dp.message_handler(commands= 'start', state='*')
async def start_commands(message: types.Message, state: FSMContext):
    to_pin = await bot.send_message(message.chat.id, 'Ознакомьтесь <b>более подробно</b> с возможностями Telegram-бота <b>ScamCheck</b>.\n\n'
                                                     '🔎<b>Справочник финансовых организаций - все организации</b>, которые предоставляют'
                                                     ' финансовые услуги, — банки, небанковские кредитные организации, негосударственные'
                                                     ' пенсионные фонды, страховые организации, управляющие компании, профессиональные'
                                                     ' участники рынка ценных бумаг, операторы финансовых платформ и другие операторы'
                                                     ' финансовых услуг, микрофинансовые организации, кооперативы, ломбарды, бюро'
                                                     ' кредитных историй,  актуарии и другие — <i>должны получить разрешение Банка'
                                                     ' России о допуске на финансовый рынок</i>. В зависимости от вида организации'
                                                     ' это может быть как лицензия, так и включение сведений в реестр или аккредитация.'
                                                     ' \n📍Для того чтобы проверить легальность работы на рынке (имеющиеся лицензии,'
                                                     ' виды деятельности), выберите в меню “<i>Справочник финансовых организаций</i>”.'
                                                     ' В полученном результате найдите интересующую вас компанию и имеющуюся по'
                                                     ' ней информацию. Если в справочнике по введенным реквизитам участника'
                                                     ' финансового рынка вы не нашли нужной информации, то высока вероятность,'
                                                     ' что компания ведет свою деятельность <i>без соответствующего разрешения</i>.'
                                                     ' Как следствие, это может повлечь за собой нарушение ваших прав. \n📍В'
                                                     ' таком случае советуем проверить компанию в “<i>Списке компаний с'
                                                     ' выявленными признаками нелегальной деятельности на финансовом рынке</i>”.'
                                                     ' Для этого выберите в меню “<i>Warning list Банка России</i>”.\n\n'
                                                     '<a href="https://www.youtube.com/watch?v=NYYfvN61mCM&t=5s&ab_channel=%D0%91%D0%B0%D0%BD%D0%BA%D0%A0%D0%BE%D1%81%D1%81%D0%B8%D0%B8">Наглядное пособие</a>'
                                                     ' как пользоваться сервисом “Справочник финансовых организаций”.\n\n'
                                                     '🆘<b>Список компаний с выявленными признаками нелегальной деятельности'
                                                     ' на финансовом рынке (Warning list Банка России)</b> - Банк России с'
                                                     ' помощью специальной системы мониторинга, а также по обращениям'
                                                     ' граждан и организаций выявляет (в том числе в Интернете) компании'
                                                     ' и проекты с признаками нелегальной деятельности. Чтобы снизить'
                                                     ' риски вовлечения граждан и организаций в незаконную деятельность,'
                                                     ' Банк России раскрывает список компаний с признаками «<i>финансовой'
                                                     ' пирамиды», нелегального кредитора, нелегального профессионального'
                                                     ' участника рынка ценных бумаг (в том числе нелегального форекс-дилера),'
                                                     'нелегального оператора инвестиционной платформы и нелегальной деятельности на страховом рынке.</i>'
                                                     ' Список не содержит сведений о физических лицах и индивидуальных'
                                                     ' предпринимателях. \n📍Для проверки наличия компании в Warning list'
                                                     ' Банка России выберите в меню “<i>Warning list Банка России</i>”.\n\n'
                                                     '📝<b>Интернет-приемная Банка России</b> - вы можете обратиться в Банк России любым удобным для Вас способом:\n'
                                                     '    - Позвонить в контактный центр☎️\n'
                                                     '    - Отправить обращение в электронном виде✉️\n'
                                                     '    - Прийти на личный прием👁️\n'
                                                     'Для получения подробной информации 📍выберите в меню “<i>Интернет-приемная</i>”.',   #Закреп
                                    parse_mode=types.ParseMode.HTML)
    await bot.pin_chat_message(chat_id=message.chat.id, message_id=to_pin['message_id'])
    # photo = open('photostart.jpg', 'rb')
    # await bot.send_photo(chat_id=message.from_user.id, photo=photo, caption='Добро пожаловать в ScamCheck - Telegram-бот, который поможет Вам сохранить финансы.\n\n'
    #                                                                                    ' 🔎Бот позволит проверить наличие компании в официальных реестрах Банка России, а также в Списке компаний с выявленными признаками нелегальной деятельности на финансовом рынке.\n\n'
    #                                                                                    ' 📝Вы также сможете ознакомиться с основными видами нелегальной деятельности на финансовом рынке и оставить жалобу на компанию через Интернет-приемную Банка России.\n\n'
    #                                                                                    ' Обезопасьте себя и своих близких, не станьте жертвой мошенников!')
    await bot.send_message(chat_id=message.from_user.id,
                           text='📖Ознакомьтесь с основными видами нелегальной деятельности на финансовом рынке',
                           reply_markup=osmenu)
    await bot.send_message(chat_id=message.from_user.id,
                           text='💰Быстрый заработок, инвестиционные проекты, пассивный доход:',
                           parse_mode='HTML',
                           reply_markup=menu1)
    await bot.send_message(chat_id=message.from_user.id,
                           text='💳 Кредитование, выдача займов, займы под залог ПТС / недвижимости:',
                           parse_mode='HTML',
                           reply_markup=menu2)
    await bot.send_message(chat_id=message.from_user.id,
                           text='📈 Брокер, форекс-дилер, доверительное управление (профессиональные участники рынка ценных бумаг): рынка ценных бумаг):',
                           parse_mode='HTML',
                           reply_markup=menu8)
    await bot.send_message(chat_id=message.from_user.id,
                           text='📄 Страхование, полисы ОСАГО::',
                           reply_markup=menu11)


@dp.callback_query_handler(text='quick_earnings', state='*')
async def quick_earnings_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await bot.send_message(chat_id=call.from_user.id,
                           text='Если вам поступило предложение инвестировать денежные средства в какую-либо компанию или проект,'
                                ' запомните - участник рынка для предоставления большинства финансовых услуг на территории'
                                ' Российской Федерации должен иметь лицензию Банка России или быть включенным в реестр регулятора.\n\n'
                                'С особой настороженностью отнеситесь к предложениям, обосновывающим свою доходность за счет деятельности,'
                                ' не регулируемой законодательством Российской Федерации. К такой деятельности может относиться обещание'
                                ' заработка за счет роста цены криптоактива (токена); ”криптовалютный арбитраж” - перепродажа монет'
                                ' по более выгодному курсу; бинарные опционы; бизнес-курсы и тренинги и т.д.\n\n'
                                'Зачастую предложенные варианты быстрого и сверхдоходного заработка обладают признаками финансовой пирамиды (схема Понци, HYIP-проект, экономические игры и т.д.).\n\n'
                                'Существует несколько общих признаков финансовых пирамид:\n\n'
                                '    ▪️обещание высокой доходности, в несколько раз превышающей рыночный уровень;\n'
                                '    ▪️гарантирование доходности (что запрещено на рынке ценных бумаг);\n'
                                '    ▪️агрессивная реклама в средствах массовой информации, сети Интернет с обещанием высокой доходности;\n'
                                '    ▪️отсутствие какой-либо информации о финансовом положении организации;\n'
                                '    ▪️выплата денежных средств за счет новых привлеченных участников;\n'
                                '    ▪️отсутствие собственных основных средств, других дорогостоящих активов;\n'
                                '    ▪️нет точного определения деятельности организации;\n'
                                '    ▪️отсутствие лицензии на осуществление деятельности на финансовом рынке / информации в реестрах Банка России.\n\n'
                                'Но наличие этих признаков финансовой пирамиды не является достаточным основанием для однозначного (безошибочного)'
                                ' вывода об отнесении той или иной организации к ней.  Для правоохранительных и надзорных органов они являются лишь'
                                ' одним из сигналов для проведения в отношении организации, которая обладает такими признаками, проверочных мероприятий.\n\n'
                                'Подробно с деятельностью финансовых пирамид также можно ознакомиться <a href="https://fincult.info/article/finansovaya-piramida-kak-ee-raspoznat/">здесь</a>.',
                           parse_mode=types.ParseMode.HTML)
    await bot.send_message(chat_id=call.from_user.id,
                           text='❗️Чтобы узнать находится ли компания в реестрах Банка России (имеет ли лицензию), <b>выберите 🔎Справочник финансовых организаций.</b>'
                                ' Также проверьте наличие сведений о компании в 🆘 <b>Списке компаний с выявленными признаками нелегальной деятельности</b>".'
                                ' Для этого выберите в меню "<b>Warning list Банка России</b>".',
                           parse_mode=types.ParseMode.HTML,
                           reply_markup=menu9)


@dp.callback_query_handler(text='osago', state='*')
async def osago_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await bot.send_message(chat_id=call.from_user.id,
                           text='Страхование является одним из ключевых финансовых институтов рыночной экономики и механизмом эффективной защиты имущественных интересов граждан, предприятий и организаций от разнообразных рисков.\n'
                                ' В настоящее время страховые услуги предоставляются в отношении более 100 видов страхования, наиболее популярными из которых являются обязательное страхование автогражданской'
                                ' ответственности (ОСАГО), страхование от несчастных случаев и болезней и страхование имущества граждан.\n'
                                ' Регулирование, контроль и надзор за деятельностью субъектов страхового дела: страховых организаций, страховых брокеров и обществ взаимного страхования осуществляет Банк России.\n\n'
                                'О том, как распознать страховых мошенников, прочитайте <a href="https://fincult.info/article/kak-raspoznat-strakhovykh-moshennikov/">тут</a>.\n '
                                '<a href="https://fincult.info/article/strakhovye-posredniki-kto-eto-i-stoit-li-k-nim-obrashchatsya/">Подробнее</a> про страховых агентов и брокеров.',
                           parse_mode=types.ParseMode.HTML)
    await bot.send_message(chat_id=call.from_user.id,
                           text='❗️Чтобы узнать находится ли компания в реестрах Банка России (имеет ли лицензию), <b>выберите 🔎Справочник финансовых организаций.</b>'
                                ' Также проверьте наличие сведений о компании в 🆘 <b>Списке компаний с выявленными признаками нелегальной деятельности</b>".'
                                ' Для этого выберите в меню "<b>Warning list Банка России</b>".',
                           parse_mode=types.ParseMode.HTML,
                           reply_markup=menu9)
@dp.callback_query_handler(text='lending_loans', state='*')
async def lending_loans_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await bot.send_message(chat_id=call.from_user.id,
                           text='Быть профессиональным кредитором, то есть выдавать кредиты и займы в денежной форме, могут'
                                ' только банки, микрофинансовые организации (МФО), кредитные потребительские кооперативы (КПК и СКПК) и ломбарды.\n\n'
                                'Если разрешения у компании (или лицензии у банка) нет, а она все равно привлекает клиентов, выдает себя за лицензированную'
                                ' и кредитует потребителей, то это нелегальный, или черный, кредитор. Такие нелегальные кредиторы могут действовать по-разному.'
                                ' Например, выдавать деньги под очень высокие проценты, но при этом не прибегать к откровенному криминалу.'
                                ' А могут использовать преступные схемы, чтобы обманом завладеть деньгами и имуществом клиентов.\n\n'
                                'Обратите внимание на сделки, совершаемые под видом финансовой аренды (лизинга). Обычно это связано'
                                ' с кредитованием под залог ПТС автомобиля. Зачастую указанные сделки могут быть притворными,'
                                ' т.е. сделками, которые совершаются с целью прикрыть другую сделку, в том числе сделку на иных условиях.\n\n'
                                'Подробно с деятельностью нелегальных кредиторов также можно ознакомиться <a href="https://fincult.info/article/kak-ne-stat-zhertvoy-chernykh-kreditorov/">здесь</a>.',
                           parse_mode=types.ParseMode.HTML)
    await bot.send_message(chat_id=call.from_user.id,
                           text='❗️Чтобы узнать находится ли компания в реестрах Банка России (имеет ли лицензию), <b>выберите 🔎Справочник финансовых организаций.</b>'
                                ' Также проверьте наличие сведений о компании в 🆘 <b>Списке компаний с выявленными признаками нелегальной деятельности</b>".'
                                ' Для этого выберите в меню "<b>Warning list Банка России</b>".',
                           parse_mode=types.ParseMode.HTML,
                           reply_markup=menu9)


@dp.callback_query_handler(text='broker', state='*')
async def lending_loans_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await bot.send_message(chat_id=call.from_user.id,
                           text='Брокеры, форекс-дилеры, доверительные управляющие (или те, кто ими прикидывается) часто ссылаются'
                                ' на то, что рынок валюты — международный. И торговать на нем можно дистанционно откуда угодно'
                                ' и через кого угодно. Это действительно так, технических ограничений нет, но есть одно «но».\n\n'
                                'Если иностранный дилер нарушит ваши права, например не будет вовремя и правильно выполнять ваши'
                                ' поручения, вам придется отстаивать свои интересы самостоятельно. Все споры будут решаться в той'
                                ' стране и по законам той страны, где зарегистрирована эта компания. Российские регулятор и суд не смогут вам помочь.\n\n'
                                'Именно поэтому по закону зарубежные компании не имеют права работать на российском финансовом рынке.'
                                ' Так что если вы увидите рекламу иностранного форекс-дилера в рунете, то он уже нарушает закон.'
                                ' Вам решать: стоит ли работать с компанией, которая нелегально привлекает клиентов.\n\n'
                                'Часто «иностранными форекс-дилерами» себя называют откровенные мошенники. Они пользуются тем,'
                                ' что обычному человеку трудно проверить зарубежную компанию. Вам дают доступ к «торговому терминалу»,'
                                ' вы совершаете «сделки», количество денег на счете растет как на дрожжах. На самом деле вы получаете'
                                ' доступ к игре, имитирующей торговлю. Ее цель — разбудить ваш азарт и заставить внести'
                                ' на счет как можно больше денег. Но как только вы захотите вывести деньги со счета,'
                                ' возникнут непреодолимые проблемы: сбой программы, технические задержки и так далее.\n\n'
                                'При этом мошенники не только не возвращают деньги, но и побуждают клиентов заплатить'
                                ' еще больше. Часто требуют комиссию за срочный вывод денег. Или объясняют, что вывести'
                                ' можно только круглую сумму, например, 10 000 $. А если у вас на счете 6700, $ то нужно'
                                ' внести еще 3300, $ и только тогда вам все вернут. Но затем появляются «новые обстоятельства», и эту сумму тоже не выплачивают.\n\n'
                                'Подробнее о том как не попасться на нелегального участника рынка ценных бумаг прочитайте <a href="https://fincult.info/article/nelegalnyy-foreks-kak-ne-popastsya-na-udochku-moshennikov/">здесь</a>.\n\n'
                                '<a href="https://fincult.info/article/chto-takoe-foreks-forex-kak-rabotaet-torgovlya-na-etom-rynke/">Гайд</a>: как начать пользоваться услугами легального форекс-дилера.',
                           parse_mode=types.ParseMode.HTML)
    await bot.send_message(chat_id=call.from_user.id,
                           text='❗️Чтобы узнать находится ли компания в реестрах Банка России (имеет ли лицензию), <b>выберите 🔎Справочник финансовых организаций.</b>'
                                ' Также проверьте наличие сведений о компании в 🆘 <b>Списке компаний с выявленными признаками нелегальной деятельности/b>".'
                                ' Для этого выберите в меню "<b>Warning list Банка России</b>".',
                           parse_mode=types.ParseMode.HTML,
                           reply_markup=menu9)


@dp.callback_query_handler(text='next_menu', state='*')
async def next_menu_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await bot.send_message(chat_id=call.from_user.id,
                           text='Для начала работы проверьте наличие компании в справочнике финансовых организаций',
                           reply_markup=menu4)


@dp.callback_query_handler(text='neleg_deyat', state='*')
async def neleg_deyat_processing(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    await state.set_state(processing_class.check.state)
    await bot.send_message(chat_id=call.from_user.id,
                           text='Поиск по названию, ИНН, адресу, сайту за весь доступный период.',
                           reply_markup=menu7)


@dp.callback_query_handler(text='name_menu2', state='*')
async def name_menu2_processing(call: types.CallbackQuery, state: FSMContext):
    await bot.send_message(chat_id=call.from_user.id,
                           text='Введите данные')
    await state.set_state(processing_class.rs.state)
    await call.answer()


# @dp.message_handler(lambda message: not message.text, state=processing_class.check.state)
# async def check_text_processing(message: types.Message, state: FSMContext):
#     return await message.answer('Это не подходящее значение, измените его')


@dp.message_handler(state=processing_class.rs.state)
async def check_org_processing(message: types.Message, state: FSMContext):
    if len(message.text) > 3:
        with open('result.json', 'rt', encoding='utf-8') as file:
            data_dict = json.loads(file.read())
        dict2 = {}
        for key in data_dict.keys():
            dict2[key] = key
        sett = set(dict2.keys())
        flaglist = []
        for i in sett:
            try:
                if len(message.text.split(' ')) == 1:
                    if message.text.lower() in i.lower():
                        flaglist.append(dict2[i])
                else:
                    text_for_search = message.text.lower().split(' ')
                    if all(j in i.lower() for j in text_for_search):
                        flaglist.append(dict2[i])
            except:
                if message.text.lower() in i.lower():
                    flaglist.append(dict2[i])
        if flaglist == []:
            await message.answer('Ничего не найдено, повторите попытку', reply_markup=menu7)
        elif len(flaglist) > 99:
            await message.answer('Выберите точнее')
            return
        else:
            flaglist.sort()
            keyb = types.InlineKeyboardMarkup()
            rubbish_list = []
            for i in flaglist:
                if data_dict[i][0] not in rubbish_list:
                    keyb.add(types.InlineKeyboardButton(text=data_dict[i][0], callback_data=f'sold-{list(data_dict.keys()).index(i)}'))
                    rubbish_list.append(data_dict[i][0])
                else:
                    pass
            keyb.add(types.InlineKeyboardButton(text='🔍 Поиск', callback_data='name_menu2'))
            await message.answer(f'Найдено компаний ({len(rubbish_list)}):', reply_markup=keyb)
            await state.update_data(flags=flaglist)
            await state.update_data(dict2=data_dict)
            await state.set_state(processing_class.result.state)
    else:
        await message.answer(text='Маленькая длина, попробуй ещё раз.')
        return


@dp.callback_query_handler(state=processing_class.result.state)
async def process_buttons(call: types.CallbackQuery, state: FSMContext):
    await call.answer()
    ind = int(call.data.split('-')[1])
    with open('result.json', 'rt', encoding='utf-8') as file:
        data_dict = json.loads(file.read())
    data_list = data_dict[list(data_dict.keys())[ind]]
    totalstr = ''
    keyb2 = types.InlineKeyboardMarkup()
    keyb2.add(types.InlineKeyboardButton(text='🔍 Поиск заново', callback_data='name_menu2'),(types.InlineKeyboardButton(text='↩️Возврат в меню выбора', callback_data='selection_menu')))
    for i, elem in enumerate(data_list):
        if elem != '' and elem != 'None':
            totalstr += f'{data_dict["Название"][i]}: <b>{elem}</b>\n'
    await bot.send_message(chat_id=call.from_user.id, text=totalstr, parse_mode='HTML', reply_markup=keyb2)


if __name__ == '__main__':
    executor.start_polling(dp,skip_updates=True)
